"""Tests for womblex.ingest.extract — extraction strategies.

Tests use real fixtures from fixtures/. OCR-dependent extractors are
tested only where PaddleOCR ONNX models are available.
"""

from pathlib import Path

import pytest

from womblex.ingest.detect import DocumentProfile, DocumentType
from womblex.ingest.extract import (
    ExtractionMetadata,
    ExtractionResult,
    PageResult,
    Position,
    _classify_native_block,
    _count_blocks_in_bbox,
    _find_native_tables,
    _normalise_bbox,
    get_extractor,
)
from womblex.ingest.strategies import (
    DocxExtractor,
    ImageExtractor,
    TextExtractor,
)
from womblex.ingest.spreadsheet import SpreadsheetExtractor


# ---------------------------------------------------------------------------
# Data model tests
# ---------------------------------------------------------------------------


class TestPosition:
    def test_fields(self) -> None:
        p = Position(x=0.1, y=0.2, width=0.5, height=0.3)
        assert p.x == 0.1
        assert p.width == 0.5

    def test_normalise_bbox(self) -> None:
        pos = _normalise_bbox((100, 200, 400, 600), 800, 1000)
        assert pos.x == pytest.approx(0.125)
        assert pos.y == pytest.approx(0.2)
        assert pos.width == pytest.approx(0.375)
        assert pos.height == pytest.approx(0.4)


class TestExtractionResult:
    def test_full_text_concatenation(self) -> None:
        result = ExtractionResult(
            pages=[
                PageResult(page_number=0, text="Hello", method="native"),
                PageResult(page_number=1, text="World", method="native"),
            ],
            method="native",
        )
        assert result.full_text == "Hello\n\nWorld"

    def test_full_text_skips_empty(self) -> None:
        result = ExtractionResult(
            pages=[
                PageResult(page_number=0, text="Hello", method="native"),
                PageResult(page_number=1, text="", method="native"),
                PageResult(page_number=2, text="End", method="native"),
            ],
            method="native",
        )
        assert result.full_text == "Hello\n\nEnd"

    def test_page_count(self) -> None:
        result = ExtractionResult(
            pages=[PageResult(page_number=i, text=f"p{i}", method="native") for i in range(5)],
            method="native",
        )
        assert result.page_count == 5

    def test_empty_result(self) -> None:
        result = ExtractionResult()
        assert result.full_text == ""
        assert result.page_count == 0
        assert result.error is None
        assert result.tables == []
        assert result.forms == []
        assert result.images == []
        assert result.text_blocks == []

    def test_structured_fields(self) -> None:
        # The legacy structured fields (tables/forms/images/text_blocks) are
        # now read-only derived views over ExtractionResult.elements. Build
        # an element stream and confirm each view projects correctly.
        from womblex.ingest.elements import Cell, Element, FieldEntry

        pos = Position(x=0.0, y=0.0, width=1.0, height=1.0)
        elements = [
            Element(
                order=0, kind="paragraph", extractor="native_text",
                bbox=pos, text="Test", confidence=0.9,
            ),
            Element(
                order=1, kind="table", extractor="native_text",
                bbox=pos, confidence=0.8,
                cells=[
                    Cell(row=0, col=0, value="A"),
                    Cell(row=0, col=1, value="B"),
                    Cell(row=1, col=0, value="1"),
                    Cell(row=1, col=1, value="2"),
                ],
                header_rows=[0],
            ),
            Element(
                order=2, kind="form", extractor="form_acroform",
                bbox=pos, confidence=0.9,
                fields=[FieldEntry(name="Name", value="Alice")],
            ),
            Element(
                order=3, kind="image", extractor="figure_image",
                bbox=pos, alt_text="photo", confidence=0.7,
            ),
        ]
        result = ExtractionResult(
            pages=[PageResult(page_number=0, text="Test", method="native")],
            elements=elements,
            method="native_with_structured",
        )
        assert len(result.tables) == 1
        assert result.tables[0].headers == ["A", "B"]
        assert len(result.forms) == 1
        assert result.forms[0].field_name == "Name"
        assert len(result.images) == 1
        assert len(result.text_blocks) == 1


class TestFindNativeTablesGate:
    """Cross-classifier gate inside ``_find_native_tables``.

    Real tables decompose into ≥1 PyMuPDF dict-block per row; prose-as-
    table over-claims rows by carving sub-block whitespace into pseudo-
    rows. The gate rejects candidates where block count < row count.
    """

    def test_count_blocks_in_bbox_uses_block_centre(self) -> None:
        import fitz

        doc = fitz.open()
        page = doc.new_page(width=400, height=600)
        page.insert_text((50, 50), "alpha")
        page.insert_text((50, 200), "beta")
        page.insert_text((50, 500), "gamma")

        bbox_top = fitz.Rect(0, 0, 400, 300)
        assert _count_blocks_in_bbox(page, bbox_top) == 2  # alpha, beta

        bbox_bottom = fitz.Rect(0, 400, 400, 600)
        assert _count_blocks_in_bbox(page, bbox_bottom) == 1  # gamma

        bbox_none = fitz.Rect(0, 300, 400, 400)
        assert _count_blocks_in_bbox(page, bbox_none) == 0

        doc.close()

    def test_gate_rejects_prose_as_table(self) -> None:
        import fitz

        doc = fitz.open()
        page = doc.new_page(width=595, height=842)
        # One paragraph block of prose with consistent left-indent. PyMuPDF's
        # text-strategy `find_tables` will read the whitespace pattern as
        # columnar and over-claim many rows; the gate should reject because
        # `n_blocks_in_bbox` (≈1) is much smaller than the claimed row count.
        prose = (
            "1. As you are aware, the Authority has issued this Notice.\n"
            "2. The Provider must comply with the requirements set out below.\n"
            "3. The Authority will continue to monitor compliance.\n"
            "4. Should you have any questions, contact the Director.\n"
            "5. This Notice takes effect immediately upon receipt.\n"
        )
        page.insert_textbox(fitz.Rect(50, 50, 545, 800), prose, fontsize=11)

        tables = _find_native_tables(page)
        # Any text-strategy hit on this page would be over-firing; the gate
        # should leave us with no tables.
        assert tables == []

        doc.close()


class TestClassifyNativeBlock:
    """Audit-cluster K1, K4, K5 — `_classify_native_block` kind output."""

    def test_yours_sincerely_is_no_longer_signature(self) -> None:
        # K1: closing phrase should not classify as 'signature'. Falls to paragraph.
        for phrase in ("Yours sincerely", "Yours faithfully,", "Yours truly"):
            assert _classify_native_block(phrase, max_font_size=12, is_bold=False, y_norm=0.7) == "paragraph"

    def test_top_of_page_short_text_is_header(self) -> None:
        # K4: short blocks at y_norm < 0.08 classify as header (not paragraph).
        assert _classify_native_block("ACT Government", max_font_size=12, is_bold=False, y_norm=0.04) == "header"

    def test_bottom_of_page_short_text_is_footer(self) -> None:
        assert _classify_native_block("3", max_font_size=10, is_bold=False, y_norm=0.97) == "footer"
        assert _classify_native_block("2 | P a g e", max_font_size=10, is_bold=False, y_norm=0.97) == "footer"

    def test_letter_subparagraph_is_list_item(self) -> None:
        # K5: (a) (b) markers classify as list_item.
        assert _classify_native_block("(a) Any discipline that is unreasonable", max_font_size=12, is_bold=False, y_norm=0.5) == "list_item"
        assert _classify_native_block("(b) other text here", max_font_size=12, is_bold=False, y_norm=0.5) == "list_item"

    def test_roman_subparagraph_is_list_item(self) -> None:
        assert _classify_native_block("(i) first item", max_font_size=12, is_bold=False, y_norm=0.5) == "list_item"
        assert _classify_native_block("(iii) third item", max_font_size=12, is_bold=False, y_norm=0.5) == "list_item"

    def test_numbered_subparagraph_is_list_item(self) -> None:
        assert _classify_native_block("(1) numbered sub-item", max_font_size=12, is_bold=False, y_norm=0.5) == "list_item"

    def test_bullet_is_list_item(self) -> None:
        assert _classify_native_block("• bullet point", max_font_size=12, is_bold=False, y_norm=0.5) == "list_item"
        assert _classify_native_block("- dash bullet", max_font_size=12, is_bold=False, y_norm=0.5) == "list_item"

    def test_bare_numbered_paragraph_is_NOT_list_item(self) -> None:
        # K5: deliberate scope — "1. " is ambiguous with numbered paragraphs, leave as paragraph.
        text = "1. As you are aware, Authorised Officers from the ACT Regulatory Authority have completed an investigation."
        assert _classify_native_block(text, max_font_size=12, is_bold=False, y_norm=0.5) == "paragraph"

    def test_heading_by_font_size(self) -> None:
        assert _classify_native_block("Compliance Notice", max_font_size=14, is_bold=True, y_norm=0.3) == "heading"

    def test_paragraph_fallback(self) -> None:
        assert _classify_native_block("Body text in a regulation citation.", max_font_size=11, is_bold=False, y_norm=0.5) == "paragraph"


class TestOcrRegionBlockType:
    """K9-fig — full-page OCR collapsed onto a non-text layout kind.

    A full-page scan OCR's to substantial prose but is tagged with the dominant
    region's kind. When that kind is `figure` (∉ TEXT_KINDS) the block is
    excluded from chunking and its content is silently lost. The helper promotes
    such a block to `paragraph` by text volume; page-furniture (page numbers,
    bare logos) keeps its `figure` kind. Text kinds (incl. `caption`) and
    `table` already reach the right consumer and pass through unchanged.
    """

    def test_substantial_text_promotes_figure_to_paragraph(self) -> None:
        from womblex.ingest.strategies_scanned import _ocr_region_block_type

        scan = (
            "13. Relevant extracts from Witness B's statement include the "
            "following observations made during the incident."
        )
        assert _ocr_region_block_type(scan, "figure") == "paragraph"

    def test_sparse_text_keeps_figure_kind(self) -> None:
        from womblex.ingest.strategies_scanned import _ocr_region_block_type

        for furniture in ("8lPage", "4l Page", "ACT\nGovernment\nEducation"):
            assert _ocr_region_block_type(furniture, "figure") == "figure"

    def test_text_kinds_and_tables_pass_through(self) -> None:
        from womblex.ingest.strategies_scanned import _ocr_region_block_type

        # A text kind (caption ∈ TEXT_KINDS) or table is never reclassified.
        assert _ocr_region_block_type("short", "heading") == "heading"
        assert _ocr_region_block_type("Figure 1 caption text here", "caption") == "caption"
        assert _ocr_region_block_type("a much longer block of prose here", "table") == "table"

    def test_threshold_boundary_is_inclusive(self) -> None:
        from womblex.ingest.strategies_scanned import _ocr_region_block_type

        assert _ocr_region_block_type("one two three four", "figure") == "figure"
        assert _ocr_region_block_type("one two three four five", "figure") == "paragraph"


class TestFormLabelDenylist:
    """K3 — `_looks_like_form_label` denylist for regulatory-letter prose."""

    def test_penalty_label_rejected(self) -> None:
        # Regulation citation: "Penalty: $10 000, in the case of an individual"
        from womblex.ingest.forms import _looks_like_form_label
        assert not _looks_like_form_label("Penalty")

    def test_official_banner_rejected(self) -> None:
        # Document classification banner: "OFFICIAL: Sensitive - Legislative Secrecy"
        from womblex.ingest.forms import _looks_like_form_label
        assert not _looks_like_form_label("OFFICIAL")

    def test_note_aside_rejected(self) -> None:
        from womblex.ingest.forms import _looks_like_form_label
        assert not _looks_like_form_label("Note")

    def test_real_form_label_still_accepted(self) -> None:
        from womblex.ingest.forms import _looks_like_form_label
        assert _looks_like_form_label("Notification Number")
        assert _looks_like_form_label("Date generated")
        assert _looks_like_form_label("Approved provider name")


class TestYoloLabelMapSelection:
    """K7(b) — DocLayNet vs COCO label-map selection from loaded class names."""

    def test_doclaynet_selected_by_section_header(self) -> None:
        from womblex.ingest.paddle_ocr import _YOLO_DOCLAYNET_LABEL_MAP, _select_label_map
        m, t = _select_label_map({0: "Section-header", 1: "Text", 2: "Title"})
        assert t == "doclaynet"
        assert m is _YOLO_DOCLAYNET_LABEL_MAP

    def test_doclaynet_selected_by_page_footer(self) -> None:
        from womblex.ingest.paddle_ocr import _select_label_map
        _, t = _select_label_map({0: "Page-footer", 1: "Text"})
        assert t == "doclaynet"

    def test_coco_fallback_for_unknown_classes(self) -> None:
        from womblex.ingest.paddle_ocr import _YOLO_COCO_LABEL_MAP, _select_label_map
        m, t = _select_label_map({0: "person", 1: "book", 2: "tv"})
        assert t == "coco"
        assert m is _YOLO_COCO_LABEL_MAP

    def test_unknown_class_defaults_to_paragraph(self) -> None:
        # K7(a) invariant preserved: unknown classes default to paragraph, not figure.
        from womblex.ingest.paddle_ocr import _YOLO_COCO_LABEL_MAP, _YOLO_DOCLAYNET_LABEL_MAP
        assert _YOLO_COCO_LABEL_MAP.get("unknown_class", "paragraph") == "paragraph"
        assert _YOLO_DOCLAYNET_LABEL_MAP.get("unknown_class", "paragraph") == "paragraph"


class TestYoloDocLayNetMap:
    """K7(b) — DocLayNet class names map to expected element kinds."""

    def test_text_classes_map_to_text_kinds(self) -> None:
        from womblex.ingest.paddle_ocr import _YOLO_DOCLAYNET_LABEL_MAP
        assert _YOLO_DOCLAYNET_LABEL_MAP["Text"] == "paragraph"
        assert _YOLO_DOCLAYNET_LABEL_MAP["Title"] == "heading"
        assert _YOLO_DOCLAYNET_LABEL_MAP["Section-header"] == "heading"
        assert _YOLO_DOCLAYNET_LABEL_MAP["List-item"] == "list_item"
        assert _YOLO_DOCLAYNET_LABEL_MAP["Caption"] == "caption"

    def test_structural_classes(self) -> None:
        from womblex.ingest.paddle_ocr import _YOLO_DOCLAYNET_LABEL_MAP
        assert _YOLO_DOCLAYNET_LABEL_MAP["Page-header"] == "header"
        assert _YOLO_DOCLAYNET_LABEL_MAP["Page-footer"] == "footer"
        assert _YOLO_DOCLAYNET_LABEL_MAP["Footnote"] == "footnote"

    def test_visual_classes(self) -> None:
        from womblex.ingest.paddle_ocr import _YOLO_DOCLAYNET_LABEL_MAP
        assert _YOLO_DOCLAYNET_LABEL_MAP["Picture"] == "figure"
        assert _YOLO_DOCLAYNET_LABEL_MAP["Table"] == "table"

    def test_formula_collapses_to_paragraph(self) -> None:
        # No dedicated formula kind — text is preserved, label collapses.
        from womblex.ingest.paddle_ocr import _YOLO_DOCLAYNET_LABEL_MAP
        assert _YOLO_DOCLAYNET_LABEL_MAP["Formula"] == "paragraph"


class TestFootnoteKind:
    """K7(b) — `footnote` is a real ElementKind, in TEXT_KINDS and block-type map."""

    def test_footnote_in_text_kinds(self) -> None:
        from womblex.ingest.elements import TEXT_KINDS
        assert "footnote" in TEXT_KINDS

    def test_footnote_in_block_type_to_kind(self) -> None:
        from womblex.ingest.orchestrator import _BLOCK_TYPE_TO_KIND
        assert _BLOCK_TYPE_TO_KIND["footnote"] == "footnote"


class TestK2PrimeOcrFormBboxes:
    """K2′ — OCR form-pair extractor preserves per-region bbox."""

    def test_region_bbox_normalised_to_position(self) -> None:
        from womblex.ingest.forms import _extract_form_pairs_from_regions
        from womblex.ingest.interfaces.protocols import OCRRegionResult

        regions = [
            OCRRegionResult(
                bbox=[[100, 200], [400, 200], [400, 230], [100, 230]],
                text="Notification Number: ABC-1234",
                confidence=0.95,
            ),
        ]
        fields = _extract_form_pairs_from_regions(regions, pix_width=600, pix_height=800)
        assert len(fields) == 1
        f = fields[0]
        assert f.field_name == "Notification Number"
        assert f.value == "ABC-1234"
        # The defining K2′ check: bbox must be a real position, not (0,0,0,0).
        assert f.position.x > 0
        assert f.position.y > 0
        assert f.position.width > 0
        assert f.position.height > 0

    def test_gap_pattern_pair(self) -> None:
        from womblex.ingest.forms import _extract_form_pairs_from_regions
        from womblex.ingest.interfaces.protocols import OCRRegionResult

        regions = [
            OCRRegionResult(
                bbox=[[50, 100], [550, 100], [550, 130], [50, 130]],
                text="Provider Name    Wonderschool",
                confidence=0.92,
            ),
        ]
        fields = _extract_form_pairs_from_regions(regions, pix_width=600, pix_height=800)
        assert len(fields) == 1
        assert fields[0].field_name == "Provider Name"
        assert fields[0].value == "Wonderschool"

    def test_denylist_still_applies(self) -> None:
        # K3 label denylist must be respected in the region-based path too.
        from womblex.ingest.forms import _extract_form_pairs_from_regions
        from womblex.ingest.interfaces.protocols import OCRRegionResult

        regions = [
            OCRRegionResult(
                bbox=[[50, 100], [400, 100], [400, 130], [50, 130]],
                text="OFFICIAL: Sensitive",
                confidence=0.90,
            ),
        ]
        fields = _extract_form_pairs_from_regions(regions, pix_width=600, pix_height=800)
        assert len(fields) == 0

    def test_empty_dims_returns_empty(self) -> None:
        from womblex.ingest.forms import _extract_form_pairs_from_regions
        assert _extract_form_pairs_from_regions([], 0, 0) == []


class TestPageBreakEmission:
    """K8 — orchestrator emits kind='page_break' between consecutive pages."""

    def test_three_page_pdf_emits_two_page_breaks(self, tmp_path: Path) -> None:
        import fitz
        from womblex.ingest.orchestrator import extract_with_plan
        from womblex.ingest.page_profile import profile_pages
        from womblex.ingest.detect import DocumentType

        doc = fitz.open()
        for i in range(3):
            page = doc.new_page(width=595, height=842)
            page.insert_text((72, 72), f"Page {i} body text")
        pdf_path = tmp_path / "multi.pdf"
        doc.save(str(pdf_path))
        doc.close()

        doc = fitz.open(str(pdf_path))
        profiles = profile_pages(doc)
        result = extract_with_plan(doc, profiles, DocumentType.NATIVE_NARRATIVE)
        doc.close()

        page_breaks = [e for e in result.elements if e.kind == "page_break"]
        assert len(page_breaks) == 2  # N-1 breaks between N pages
        # Each page_break is associated with the page it precedes
        assert page_breaks[0].page == 1
        assert page_breaks[1].page == 2

    def test_single_page_pdf_emits_no_page_breaks(self, tmp_path: Path) -> None:
        import fitz
        from womblex.ingest.orchestrator import extract_with_plan
        from womblex.ingest.page_profile import profile_pages
        from womblex.ingest.detect import DocumentType

        doc = fitz.open()
        page = doc.new_page(width=595, height=842)
        page.insert_text((72, 72), "single page")
        pdf_path = tmp_path / "single.pdf"
        doc.save(str(pdf_path))
        doc.close()

        doc = fitz.open(str(pdf_path))
        profiles = profile_pages(doc)
        result = extract_with_plan(doc, profiles, DocumentType.NATIVE_NARRATIVE)
        doc.close()

        assert not any(e.kind == "page_break" for e in result.elements)


class TestExtractionMetadata:
    def test_fields(self) -> None:
        m = ExtractionMetadata(
            extraction_strategy="native_narrative",
            confidence=0.95,
            processing_time=1.2,
            page_count=3,
            text_coverage=0.9,
        )
        assert m.extraction_strategy == "native_narrative"
        assert m.preprocessing_steps == []
        assert m.content_mix == {}


# ---------------------------------------------------------------------------
# get_extractor
# ---------------------------------------------------------------------------


class TestGetExtractor:
    """`get_extractor` is the legacy non-PDF dispatch. Native and scanned PDFs
    route through `extract_pdf_with_plan` (orchestrator) — see
    test_orchestrator-style assertions in test_pipeline.py / test_integration.py
    for that path."""

    def _make_profile(self, doc_type: DocumentType) -> DocumentProfile:
        return DocumentProfile(
            doc_type=doc_type,
            page_count=1,
            has_text_layer=True,
            text_coverage=1.0,
            has_images=False,
            has_tables=False,
            has_handwriting_signals=False,
            ocr_confidence=None,
            glyph_regularity=None,
            stroke_consistency=None,
            confidence=0.9,
        )

    def test_image_returns_correct_extractor(self) -> None:
        ext = get_extractor(self._make_profile(DocumentType.IMAGE))
        assert isinstance(ext, ImageExtractor)

    def test_spreadsheet_returns_correct_extractor(self) -> None:
        ext = get_extractor(self._make_profile(DocumentType.SPREADSHEET))
        assert isinstance(ext, SpreadsheetExtractor)

    def test_docx_returns_correct_extractor(self) -> None:
        ext = get_extractor(self._make_profile(DocumentType.DOCX))
        assert isinstance(ext, DocxExtractor)

    def test_text_returns_correct_extractor(self) -> None:
        ext = get_extractor(self._make_profile(DocumentType.TEXT))
        assert isinstance(ext, TextExtractor)

    def test_pdf_type_raises(self) -> None:
        with pytest.raises(ValueError, match="only handles non-PDF"):
            get_extractor(self._make_profile(DocumentType.NATIVE_NARRATIVE))


# ---------------------------------------------------------------------------
# SpreadsheetExtractor (uses real fixture spreadsheets)
# ---------------------------------------------------------------------------


class TestSpreadsheetExtractor:
    """One ExtractionResult per workbook with cell-grained elements.

    The previous one-result-per-row shape was removed; a workbook now
    yields a single result whose ``elements`` are sheet_meta + sheet_cell
    in workbook order.
    """

    def test_extracts_real_csv(self, spreadsheet_dir: Path) -> None:
        csv_path = spreadsheet_dir / "Approved-providers-au-export_20260204.csv"
        if not csv_path.exists():
            pytest.skip("CSV fixture not available")

        ext = SpreadsheetExtractor()
        result = ext.extract_path(csv_path)

        assert result.method == "spreadsheet"
        assert result.error is None
        assert result.metadata is not None
        kinds = [e.kind for e in result.elements]
        assert kinds.count("sheet_meta") >= 1
        assert kinds.count("sheet_cell") >= 1

    def test_extracts_real_xlsx(self, spreadsheet_dir: Path) -> None:
        xlsx_path = spreadsheet_dir / "mso-statistics-sept-qtr-2025.xlsx"
        if not xlsx_path.exists():
            pytest.skip("Excel fixture not available")

        ext = SpreadsheetExtractor()
        result = ext.extract_path(xlsx_path)

        assert result.method == "spreadsheet"
        assert result.metadata is not None
        # Multi-sheet workbook: one sheet_meta per sheet.
        n_sheet_meta = sum(1 for e in result.elements if e.kind == "sheet_meta")
        assert n_sheet_meta >= 1

    def test_handles_missing_file(self, tmp_path: Path) -> None:
        ext = SpreadsheetExtractor()
        result = ext.extract_path(tmp_path / "missing.csv")
        assert result.error is not None
        assert result.elements == []

    def test_xlsx_title_preamble_split_from_header(self, tmp_path: Path) -> None:
        """Export-product shape (e.g. AusTender): title + blank row above the header.

        The real header row must be the one marked is_header — pandas'
        fabricated ``Unnamed: N`` names must never appear as cell values.
        """
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Contract Notice Export"
        ws.append(["Contract Notice Export"])
        ws.append([])
        headers = ["Agency", "CN ID", "Publish Date", "Value", "Description", "Supplier Name"]
        ws.append(headers)
        ws.append(["Example Agency", "CN0000001", "2026-01-05 09:00:00", "12345.67",
                   "Office fitout", "Example Supplier Pty Ltd"])
        ws.append(["Example Agency", "CN0000002", "2026-01-06 10:00:00", "99000",
                   "ICT services", "Another Supplier Pty Ltd"])
        path = tmp_path / "AusTenderContractNoticeExport_20260101_000000.xlsx"
        wb.save(path)

        result = SpreadsheetExtractor().extract_path(path)
        assert result.error is None

        header_cells = [
            e.value for e in result.elements
            if e.kind == "sheet_cell" and (e.meta or {}).get("is_header")
        ]
        assert header_cells == headers
        assert not any("Unnamed" in (e.value or "") for e in result.elements)

        sheet_meta = next(e for e in result.elements if e.kind == "sheet_meta")
        assert sheet_meta.meta is not None
        assert sheet_meta.meta["preamble"] == "Contract Notice Export"

        data_cells = [
            e.value for e in result.elements
            if e.kind == "sheet_cell" and e.row == 1
        ]
        assert data_cells[0] == "Example Agency"
        assert "Example Supplier Pty Ltd" in data_cells

    def test_xlsx_multicell_metadata_preamble(self, tmp_path: Path) -> None:
        """Metadata blocks above the header span 2 cells but are not the header.

        A naive "first row with >=2 non-empty cells" rule would pick the
        ``Agency: | ...`` row; the width-ratio rule keeps scanning until a
        row matching the table's real width.
        """
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Quarterly Statistics Report"])
        ws.append(["Agency:", "Example Commission"])
        ws.append(["Generated:", "2026-06-11"])
        ws.append([])
        headers = ["Region", "Quarter", "Applications", "Approvals", "Refusals", "Pending"]
        ws.append(headers)
        ws.append(["NSW", "Q3", "120", "100", "15", "5"])
        path = tmp_path / "stats.xlsx"
        wb.save(path)

        result = SpreadsheetExtractor().extract_path(path)
        assert result.error is None

        header_cells = [
            e.value for e in result.elements
            if e.kind == "sheet_cell" and (e.meta or {}).get("is_header")
        ]
        assert header_cells == headers
        sheet_meta = next(e for e in result.elements if e.kind == "sheet_meta")
        assert sheet_meta.meta is not None
        preamble = sheet_meta.meta["preamble"]
        assert "Quarterly Statistics Report" in preamble
        assert "Example Commission" in preamble
        assert "2026-06-11" in preamble

    def test_ragged_csv_with_title_row(self, tmp_path: Path) -> None:
        """A one-field title row above a wide CSV header must not fail the read."""
        path = tmp_path / "export.csv"
        path.write_text(
            "Contract Notice Export\n"
            "\n"
            "Agency,CN ID,Value,Description\n"
            "Example Agency,CN0000001,12345.67,Office fitout\n",
            encoding="utf-8",
        )

        result = SpreadsheetExtractor().extract_path(path)
        assert result.error is None

        header_cells = [
            e.value for e in result.elements
            if e.kind == "sheet_cell" and (e.meta or {}).get("is_header")
        ]
        assert header_cells == ["Agency", "CN ID", "Value", "Description"]
        sheet_meta = next(e for e in result.elements if e.kind == "sheet_meta")
        assert (sheet_meta.meta or {}).get("preamble") == "Contract Notice Export"

    def test_two_column_key_value_sheet_keeps_row0_header(self, tmp_path: Path) -> None:
        """Uniformly narrow sheets are not mistaken for preamble + header."""
        path = tmp_path / "kv.csv"
        path.write_text(
            "Field,Value\nLicence number,L1234\nStatus,Active\n",
            encoding="utf-8",
        )

        result = SpreadsheetExtractor().extract_path(path)
        assert result.error is None

        header_cells = [
            e.value for e in result.elements
            if e.kind == "sheet_cell" and (e.meta or {}).get("is_header")
        ]
        assert header_cells == ["Field", "Value"]
        sheet_meta = next(e for e in result.elements if e.kind == "sheet_meta")
        assert "preamble" not in (sheet_meta.meta or {})

    def test_title_wider_than_table_not_header(self, tmp_path: Path) -> None:
        """A title row spanning more cells than the table itself stays preamble.

        Run-scoring catches what a pure width rule cannot: the title's run
        is broken by the blank row below it, while the real header is
        followed by its data body.
        """
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Licence", "Register", "Extract"])  # 3-cell title
        ws.append([])
        ws.append(["Field", "Value"])  # real 2-col header
        ws.append(["Licence number", "L1234"])
        ws.append(["Status", "Active"])
        path = tmp_path / "register.xlsx"
        wb.save(path)

        result = SpreadsheetExtractor().extract_path(path)
        assert result.error is None

        header_cells = [
            e.value for e in result.elements
            if e.kind == "sheet_cell" and (e.meta or {}).get("is_header")
        ]
        assert header_cells == ["Field", "Value"]
        sheet_meta = next(e for e in result.elements if e.kind == "sheet_meta")
        assert sheet_meta.meta is not None
        assert "Licence" in sheet_meta.meta["preamble"]

    def test_sub_header_row_does_not_shift_header(self, tmp_path: Path) -> None:
        """A single-cell section row directly under the header is neutral."""
        path = tmp_path / "sectioned.csv"
        path.write_text(
            "Name,Code,Status\n"
            "Section A,,\n"
            "Alpha,A1,Active\n"
            "Beta,B2,Active\n",
            encoding="utf-8",
        )

        result = SpreadsheetExtractor().extract_path(path)
        assert result.error is None

        header_cells = [
            e.value for e in result.elements
            if e.kind == "sheet_cell" and (e.meta or {}).get("is_header")
        ]
        assert header_cells == ["Name", "Code", "Status"]

    def test_csv_without_preamble_unchanged(self, tmp_path: Path) -> None:
        """A plain header-first CSV keeps its row-0 header behaviour."""
        path = tmp_path / "plain.csv"
        path.write_text("Name,Code\nAlpha,A1\nBeta,B2\n", encoding="utf-8")

        result = SpreadsheetExtractor().extract_path(path)
        assert result.error is None

        header_cells = [
            e.value for e in result.elements
            if e.kind == "sheet_cell" and (e.meta or {}).get("is_header")
        ]
        assert header_cells == ["Name", "Code"]
        sheet_meta = next(e for e in result.elements if e.kind == "sheet_meta")
        assert "preamble" not in (sheet_meta.meta or {})
        rows = {
            (e.row, e.col): e.value
            for e in result.elements if e.kind == "sheet_cell"
        }
        assert rows[(1, 0)] == "Alpha"
        assert rows[(2, 1)] == "B2"


# ---------------------------------------------------------------------------
# DocxExtractor
# ---------------------------------------------------------------------------


class TestDocxExtractor:
    def test_handles_missing_docx_library(self, tmp_path: Path) -> None:
        ext = DocxExtractor()
        result = ext.extract_path(tmp_path / "test.docx")
        # python-docx not installed in test env
        assert result.error is not None
        assert "docx" in result.error.lower()
