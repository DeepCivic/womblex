"""Spreadsheet extraction — one ExtractionResult per workbook.

Cells are the element grain. Every non-empty cell becomes one
``Element`` of kind ``sheet_cell``; each sheet emits a leading
``sheet_meta`` element carrying its index and dimensions.

This is the shape change relative to the previous extractor, which
produced one ExtractionResult per logical row. That shape forced
spreadsheets to masquerade as documents and made cross-format queries
awkward; the element-stream model treats cells natively.

Source values are verbatim: pandas reads with ``dtype=str``, so
"1,234" stays "1,234". That read is authoritative for ``value`` and is
never overridden here.

For XLSX a second, read-only openpyxl pass supplies the two facts pandas
discards: the cell's ``number_format`` and whether the stored value was
numeric (``value_type``). Both are cell *metadata*, not content — the
string on ``value`` is unchanged either way. They matter because a
register's money column is often identifiable only from its format:
a GrantConnect award export carries ``$#,##0.00`` on ~49,000 cells whose
text is a bare ``50000``, with no currency symbol anywhere in the sheet.
Dropping the format discards the only unambiguous currency marker in the
file. Only non-``General`` formats are retained, which keeps the lookup
small (most cells are ``General``). CSV has no cell formats, so those
sheets keep ``number_format`` unset.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    import pandas as pd

    from womblex.ingest.detect import DocumentProfile

from womblex.ingest.detect import SheetInfo
from womblex.ingest.elements import Element
from womblex.ingest.extract import (
    ExtractionMetadata,
    ExtractionResult,
    PageResult,
)

logger = logging.getLogger(__name__)

# Rows scanned for the real header when a sheet opens with title /
# blank / metadata rows. Wide enough to clear a long leading key/value
# block (e.g. the ~20-row GrantConnect "Criteria Summary" above its real
# 32-column header) as well as the short AusTender title/blank preamble.
_HEADER_SCAN_ROWS = 40

# Extra rows past the scan window used to score how table-like the rows
# below a header candidate are. With the scan window above this gives a
# 60-row horizon — enough for a real header's data body to out-score a
# narrow metadata run that precedes it.
_HEADER_LOOKAHEAD_ROWS = 20

# Fallback discriminator when run-scoring cannot decide: a header
# candidate must span at least this fraction of the widest row in the
# scan window.
_HEADER_WIDTH_RATIO = 0.6


def read_csv_raw(path: Path, *, nrows: int | None = None) -> pd.DataFrame:
    """Read a CSV with no header inference, tolerating ragged leading rows.

    ``pd.read_csv(header=None)`` infers the column count from the first
    row, so a one-field title row above a wide header makes the whole
    file fail ("Expected 1 fields, saw N"). Sniffing the true maximum
    field count and passing explicit column names lets export-product
    CSVs with preamble rows parse; short rows are padded with ``""``.
    With ``nrows``, the sniff is capped at the same row count — pandas
    parses no further, so later rows cannot affect the read.
    """
    import csv
    from itertools import islice

    import pandas as pd

    with open(path, newline="", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        rows = islice(reader, nrows) if nrows is not None else reader
        max_fields = max((len(row) for row in rows), default=0)
    df = pd.read_csv(
        path, dtype=str, keep_default_na=False, header=None,
        names=range(max_fields), nrows=nrows,
    )
    return df.fillna("")


def split_preamble(df_raw: pd.DataFrame) -> tuple[list[str], pd.DataFrame]:
    """Split leading title/metadata rows from a ``header=None`` sheet read.

    Export products put title rows, generated-date lines or ``key: value``
    metadata blocks above the real header; parsing those with pandas'
    default ``header=0`` turns the first row into the header and
    fabricates ``Unnamed: N`` column names — values that never existed in
    the source. Reading with ``header=None`` and splitting here keeps
    everything verbatim.

    The header is the candidate row (>= 2 non-empty cells in the scan
    window) that maximises ``width * run`` — the breadth of the row times
    the length of the run of table-consistent rows below it (a following
    row is consistent when it has >= 2 non-empty cells and is no wider
    than the candidate, +1 column of slack). Scoring on breadth-times-depth
    is what lets a wide real header out-rank a *longer* but narrow run of
    ``key | value`` metadata above it: the GrantConnect "Criteria Summary"
    is a ~20-row block of 2-wide rows, but the 32-wide header below it
    starts a far higher-scoring body. Title and metadata rows also score
    low because a blank separator or the wider table below breaks their
    run. Ties on score prefer the wider candidate. When no candidate has
    any consistent following row (e.g. a header-only sheet), falls back to
    the width-ratio rule, and to row 0 for single-column sheets — so
    headerless and uniformly narrow layouts keep their old behaviour.
    """
    n = len(df_raw)
    if n == 0:
        return [], df_raw

    scan = min(n, _HEADER_SCAN_ROWS)
    horizon = min(n, _HEADER_SCAN_ROWS + _HEADER_LOOKAHEAD_ROWS)
    widths = [
        sum(1 for v in df_raw.iloc[i] if str(v).strip())
        for i in range(horizon)
    ]

    def _run_length(idx: int) -> int:
        cap = widths[idx] + 1
        run = 0
        for j in range(idx + 1, horizon):
            w = widths[j]
            if w == 1:
                # Single-cell rows (sub-headers, section notes) neither
                # extend nor break a table run.
                continue
            if w == 0 or w > cap:
                break
            run += 1
        return run

    header_idx = 0
    best_score = 0
    for i in range(scan):
        if widths[i] < 2:
            continue
        score = _run_length(i) * widths[i]
        if score > best_score or (score == best_score and score > 0 and widths[i] > widths[header_idx]):
            header_idx, best_score = i, score

    if best_score == 0:
        # Nothing below any candidate looks like a table body — fall back
        # to width: first row reaching the ratio of the window's widest.
        max_width = max(widths[:scan])
        if max_width >= 2:
            threshold = max(2.0, _HEADER_WIDTH_RATIO * max_width)
            header_idx = next(
                (i for i, w in enumerate(widths[:scan]) if w >= threshold), 0,
            )

    preamble = [
        str(v).strip()
        for i in range(header_idx)
        for v in df_raw.iloc[i]
        if str(v).strip()
    ]
    df = df_raw.iloc[header_idx + 1:].reset_index(drop=True)
    df.columns = [str(v) for v in df_raw.iloc[header_idx]]
    return preamble, df


def _classify_sheet(name: str, df: pd.DataFrame) -> SheetInfo:
    """Classify a sheet's structure for detection-time metadata only.

    The new extractor emits every cell regardless of sheet shape, so
    classification no longer routes extraction. It remains useful as a
    downstream hint on ``DocumentProfile.sheet_meta``.

    Categories: ``narrative`` (single column or very long cells),
    ``glossary`` (two columns at 50–500 rows), ``key_value`` (two columns
    under 50 rows), ``data`` (everything else).
    """
    rows, cols = len(df), len(df.columns)
    if cols == 0:
        return SheetInfo(
            name=name, sheet_type="data", row_count=rows, col_count=0,
            key_column=None, has_sub_headers=False,
        )
    try:
        avg_len = float(df.apply(lambda c: c.astype(str).str.len()).stack().mean())
    except Exception:
        avg_len = 0.0

    if cols == 1 or (cols <= 3 and avg_len > 150):
        sheet_type = "narrative"
    elif cols == 2 and 50 <= rows <= 500:
        sheet_type = "glossary"
    elif cols == 2 and rows < 50:
        sheet_type = "key_value"
    else:
        sheet_type = "data"

    headers_lower = [str(c).lower() for c in df.columns]
    key_column: str | None = None
    for kw in ("id", "name", "code", "key"):
        key_column = next(
            (str(df.columns[i]) for i, h in enumerate(headers_lower) if kw in h),
            None,
        )
        if key_column:
            break
    if key_column is None:
        key_column = str(df.columns[0])

    has_sub_headers = False
    if cols > 1:
        first_nonempty = df.iloc[:, 0].astype(str).str.strip().str.len() > 0
        others_empty = ~df.iloc[:, 1:].astype(str).apply(
            lambda c: c.str.strip().str.len() > 0
        ).any(axis=1)
        has_sub_headers = int((first_nonempty & others_empty).sum()) > 2

    return SheetInfo(
        name=name, sheet_type=sheet_type, row_count=rows, col_count=cols,
        key_column=key_column, has_sub_headers=has_sub_headers,
    )


class SpreadsheetExtractor:
    """Read an XLSX / CSV workbook into a single element stream.

    Element order is monotonic across the workbook: each sheet's
    sheet_meta element is followed by its cell elements in (row, col)
    order. Sheets are emitted in their natural workbook index order.
    """

    def __init__(self, profile: DocumentProfile | None = None) -> None:
        # ``profile`` is accepted for parity with previous call sites
        # (the document profiler may pre-classify sheets). The new
        # extractor does not use the classification — every cell becomes
        # an element regardless of sheet shape.
        self.profile = profile

    def extract_path(self, path: Path) -> ExtractionResult:
        import pandas as pd

        elements: list[Element] = []
        order = 0
        suffix = path.suffix.lower()
        try:
            if suffix == ".csv":
                df_raw = read_csv_raw(path)
                preamble, df = split_preamble(df_raw)
                order = _emit_sheet(
                    elements, order, sheet_name="default", sheet_index=0,
                    df=df, preamble=preamble,
                )
            else:
                xl = pd.ExcelFile(str(path))
                cell_meta = _read_cell_metadata(path)
                for sheet_idx, name in enumerate(xl.sheet_names):
                    df_raw = xl.parse(name, dtype=str, keep_default_na=False, header=None)
                    preamble, df = split_preamble(df_raw)
                    # split_preamble drops rows 0..header_idx inclusive, so the
                    # header's index in the original sheet is recoverable from
                    # the row counts — no signature change needed to align the
                    # openpyxl lookup with the emitted grid.
                    header_idx = len(df_raw) - len(df) - 1
                    order = _emit_sheet(
                        elements, order,
                        sheet_name=str(name), sheet_index=sheet_idx,
                        df=df, preamble=preamble,
                        cell_meta=cell_meta.get(str(name)),
                        row_offset=max(header_idx, 0),
                    )
        except Exception as e:
            return _spreadsheet_error(path.stem, f"Failed to read spreadsheet: {e}")

        # Page text: one line per cell, prefixed with sheet/row/col.
        # Downstream consumers reading ``full_text`` see a flattened view;
        # the canonical structure is on ``elements``.
        page_text = "\n".join(
            f"[{e.sheet}!{e.row},{e.col}] {e.value}"
            for e in elements
            if e.kind == "sheet_cell" and e.value
        )

        return ExtractionResult(
            pages=[PageResult(page_number=0, text=page_text, method="spreadsheet")],
            elements=elements,
            method="spreadsheet",
            document_id=path.stem,
            metadata=ExtractionMetadata(
                extraction_strategy="spreadsheet",
                confidence=0.95,
                processing_time=0.0,
                page_count=1,
                text_coverage=1.0 if page_text else 0.0,
            ),
        )


def _value_type_for(value: object) -> str:
    """Classify a stored spreadsheet value. A hint only — never a coercion."""
    import datetime as _dt

    if isinstance(value, bool):
        return "bool"
    if isinstance(value, (_dt.datetime, _dt.date, _dt.time)):
        return "date"
    if isinstance(value, (int, float)):
        return "numeric"
    return "text"


def _read_cell_metadata(
    path: Path,
) -> dict[str, dict[tuple[int, int], tuple[str, str]]]:
    """Map ``sheet -> {(row, col): (number_format, value_type)}`` for an XLSX.

    Row/column indices are zero-based over the raw sheet, matching pandas'
    ``header=None`` read so callers can offset into the emitted grid. Only
    cells whose format is not ``General`` are recorded, so a mostly-unformatted
    workbook costs almost nothing.

    Best-effort: any failure yields an empty map and extraction proceeds with
    the fields unset, exactly as before this pass existed.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover - openpyxl is a core dependency
        logger.warning("openpyxl unavailable; cell formats not preserved")
        return {}

    out: dict[str, dict[tuple[int, int], tuple[str, str]]] = {}
    try:
        wb = load_workbook(str(path), read_only=True, data_only=True)
    except Exception as e:
        logger.warning("Could not read cell formats from %s: %s", path.name, e)
        return {}

    try:
        for ws in wb.worksheets:
            sheet_map: dict[tuple[int, int], tuple[str, str]] = {}
            for r, row in enumerate(ws.iter_rows()):
                for cell in row:
                    if cell.value is None:
                        continue
                    fmt = cell.number_format or ""
                    vtype = _value_type_for(cell.value)
                    if fmt in ("", "General") and vtype == "text":
                        continue
                    sheet_map[(r, cell.column - 1)] = (fmt, vtype)
            if sheet_map:
                out[ws.title] = sheet_map
    except Exception as e:
        logger.warning("Cell-format pass failed on %s: %s", path.name, e)
        return out
    finally:
        wb.close()
    return out


def _emit_sheet(
    elements: list[Element],
    start_order: int,
    *,
    sheet_name: str,
    sheet_index: int,
    df: pd.DataFrame,
    preamble: list[str] | None = None,
    cell_meta: dict[tuple[int, int], tuple[str, str]] | None = None,
    row_offset: int = 0,
) -> int:
    """Append sheet_meta + sheet_cell elements for one sheet. Return next order.

    Title/blank rows split off above the header land verbatim on the
    sheet_meta element (``meta["preamble"]``), keeping the row-0-is-header
    contract of the cell grid intact for downstream table views.

    ``cell_meta`` carries per-cell ``(number_format, value_type)`` from the
    openpyxl pass, keyed by raw-sheet coordinates; ``row_offset`` is the raw
    index of the header row, so emitted ``row`` values map back onto it.
    """
    order = start_order
    n_rows, n_cols = len(df), len(df.columns)
    meta = {
        "sheet_index": str(sheet_index),
        "rows": str(n_rows),
        "cols": str(n_cols),
    }
    if preamble:
        meta["preamble"] = "\n".join(preamble)
    elements.append(Element(
        order=order, kind="sheet_meta", extractor="xlsx",
        sheet=sheet_name, confidence=1.0,
        meta=meta,
    ))
    order += 1

    def _lookup(row_idx: int, col_idx: int) -> tuple[str | None, str]:
        if not cell_meta:
            return None, "text"
        found = cell_meta.get((row_offset + row_idx, col_idx))
        if found is None:
            return None, "text"
        fmt, vtype = found
        return (fmt or None), vtype

    # Header row at row=0. Headers are labels, so their own format is not
    # meaningful — but the label itself is what identifies a column
    # downstream, so it is emitted unchanged.
    for col_idx, header in enumerate(df.columns):
        text = str(header)
        if not text.strip():
            continue
        elements.append(Element(
            order=order, kind="sheet_cell", extractor="xlsx",
            sheet=sheet_name, row=0, col=col_idx,
            value=text, value_type="text", confidence=1.0,
            meta={"is_header": "true"},
        ))
        order += 1

    # Data rows at row=1..n
    for row_idx, (_, row) in enumerate(df.iterrows(), start=1):
        for col_idx, val in enumerate(row.values):
            text = str(val)
            if not text.strip():
                continue
            number_format, value_type = _lookup(row_idx, col_idx)
            elements.append(Element(
                order=order, kind="sheet_cell", extractor="xlsx",
                sheet=sheet_name, row=row_idx, col=col_idx,
                value=text, value_type=value_type, confidence=1.0,
                number_format=number_format,
            ))
            order += 1

    return order


def _spreadsheet_error(stem: str, msg: str) -> ExtractionResult:
    return ExtractionResult(
        pages=[], method="spreadsheet", error=msg, document_id=stem,
        metadata=ExtractionMetadata(
            extraction_strategy="spreadsheet",
            confidence=0.0, processing_time=0.0,
            page_count=0, text_coverage=0.0,
        ),
    )
